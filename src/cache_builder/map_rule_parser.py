"""
Parser for Map Rule Specification (reference document).

Not needed at runtime but parsed for audit trail completeness.
Maps function names (COPY, DECODE, USUBJID) to their descriptions.
"""

from __future__ import annotations
from pathlib import Path

import openpyxl

from config.settings import AZ_MAP_RULE_FILE, EXCEL
from src.cache_builder.models import MapRule
from src.utils.logging_config import get_logger

logger = get_logger(__name__)


def parse_map_rules(filepath: Path | None = None) -> list[MapRule]:
    """
    Parse Map Rule Specification file.

    Column positions from inspection:
        0: Mapping Function (may contain multi-line text)
        1: Description
        2: Code
        3: Map Definition

    Returns list of MapRule objects.
    """
    filepath = filepath or AZ_MAP_RULE_FILE

    if not filepath.exists():
        logger.warning("Map Rule file not found — skipping", path=str(filepath))
        return []

    logger.info("Parsing Map Rule Specification")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    try:
        if EXCEL.map_rule_sheet not in wb.sheetnames:
            logger.warning(
                "Map Rule sheet not found",
                expected=EXCEL.map_rule_sheet,
                available=wb.sheetnames,
            )
            return []

        ws = wb[EXCEL.map_rule_sheet]
        rules: list[MapRule] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            func_raw = str(row[0]).strip()
            if not func_raw:
                continue

            # Function name can be multi-line; take first line only
            func_name = func_raw.split("\n")[0].strip()
            # Remove trailing whitespace artifacts
            func_name = func_name.rstrip("\xa0").strip()

            rules.append(MapRule(
                function_name=func_name,
                description=str(row[1]).strip().rstrip("\xa0") if len(row) > 1 and row[1] else "",
                code=str(row[2]).strip().rstrip("\xa0") if len(row) > 2 and row[2] else "",
                map_definition=str(row[3]).strip().rstrip("\xa0") if len(row) > 3 and row[3] else "",
            ))

    finally:
        wb.close()

    logger.info("Map Rules parsed", count=len(rules))
    return rules