"""
Parser for the AZ RAW-to-SDTM Specification Template Excel file.

Handles:
- Two-row header structure (row 1 = super-header, row 2 = column names)
- Duplicate column names across source/target sides
- Grouped rows where Library and Dataset inherit from above
- Multiple sheets with different column counts (main vs. BMRES2OM vs. PATHBIO)
- Empty rows, trailing whitespace, inconsistent formatting

Output:
- Complete list of SpecMappingEntry objects
- SpecModuleIndex for quick module → domain lookups
"""

from __future__ import annotations
from pathlib import Path
from collections import Counter

import openpyxl

from config.settings import AZ_SPEC_FILE, EXCEL
from src.cache_builder.models import SpecMappingEntry, SpecModuleIndex
from src.utils.text_normalizer import normalize_label_for_lookup, normalize_module_name
from src.utils.logging_config import get_logger

logger = get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Column Layout Detection
# ─────────────────────────────────────────────────────────────────────────────

class _ColumnLayout:
    """
    Detected column positions for a spec sheet.

    Solves the duplicate-column-name problem by finding the SDTM boundary
    (where columns transition from source to target fields).
    """

    __slots__ = (
        "src_library", "src_dataset", "src_variable", "src_label",
        "map_definition", "map_rule", "map_mode", "map_order",
        "sdtm_dataset", "sdtm_variable", "sdtm_label", "sdtm_core",
        "total_cols",
    )

    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)

    def __repr__(self) -> str:
        return (
            f"_ColumnLayout(src_dataset={self.src_dataset}, src_label={self.src_label}, "
            f"sdtm_dataset={self.sdtm_dataset}, sdtm_variable={self.sdtm_variable}, "
            f"sdtm_core={self.sdtm_core})"
        )


def _detect_layout(header_cells: tuple) -> _ColumnLayout | None:
    """
    Detect column layout from the actual header row values.

    Strategy:
    1. Normalize all header cell values to lowercase
    2. Find ALL positions of 'dataset', 'variable', 'label'
    3. The first occurrence of each = source; second = SDTM target
    4. 'library' is always source-only; 'core' is always target-only

    Known layouts from inspection:
    - Main sheet (12 cols): Library|Dataset|Variable|Label|MapDef|MapRule|MapMode|MapOrder|Dataset|Variable|Label|Core
    - BMRES2OM (12 cols):   Library|Dataset|Variable|Label|MapDef|MapOrder|Dataset|Variable|Label|Core|_|_
    - PATHBIO (12 cols):    Library|Dataset|Variable|Label|MapDef|MapRule|MapMode|MapOrder|Dataset|Variable|Label|Core
    """
    headers = [str(h).strip().lower() if h else "" for h in header_cells]
    total_cols = len(headers)

    def find_all(name: str) -> list[int]:
        """Find all positions of a column name."""
        return [i for i, h in enumerate(headers) if h == name]

    def find_first(name: str, start: int = 0) -> int | None:
        """Find first position of column name at or after start."""
        for i in range(start, len(headers)):
            if headers[i] == name:
                return i
        return None

    def find_any(names: list[str], start: int = 0) -> int | None:
        """Find first position matching any of the given names."""
        for i in range(start, len(headers)):
            if headers[i] in names:
                return i
        return None

    # Find key columns
    dataset_positions = find_all("dataset")
    variable_positions = find_all("variable")
    label_positions = find_all("label")

    # We need at least 2 'dataset' and 2 'variable' columns (source + target)
    if len(dataset_positions) < 2 or len(variable_positions) < 2:
        logger.warning(
            "Insufficient dataset/variable columns detected",
            dataset_positions=dataset_positions,
            variable_positions=variable_positions,
            headers=headers,
        )
        return None

    # Source side: first occurrences
    src_dataset = dataset_positions[0]
    src_variable = variable_positions[0]
    src_label = label_positions[0] if label_positions else src_variable + 1

    # Library is always column 0 (or wherever 'library' appears)
    src_library = find_first("library") or 0

    # SDTM side: second occurrences
    sdtm_dataset = dataset_positions[1]
    sdtm_variable = variable_positions[1]
    sdtm_label = label_positions[1] if len(label_positions) >= 2 else sdtm_variable + 1

    # Core: always after SDTM label
    sdtm_core = find_first("core", sdtm_label)
    if sdtm_core is None:
        # Some sheets use "Required"/"Permitted" in Core column without header "Core"
        # Try column after sdtm_label
        sdtm_core = sdtm_label + 1 if sdtm_label + 1 < total_cols else None

    # Mapping columns: between source and SDTM sides
    # These appear between src_label and sdtm_dataset
    mid_start = src_label + 1 if src_label is not None else src_variable + 2
    mid_end = sdtm_dataset

    map_definition = find_any(["map definition", "mapdefinition"], mid_start)
    if map_definition is None or map_definition >= mid_end:
        map_definition = find_any(["map definition"], 0)

    map_rule = find_any(["map rule", "maprule"], mid_start)
    if map_rule is not None and map_rule >= mid_end:
        map_rule = None

    map_mode = find_any(["map mode", "mapmode"], mid_start)
    if map_mode is not None and map_mode >= mid_end:
        map_mode = None

    map_order = find_any(["map order", "maporder"], mid_start)
    if map_order is not None and map_order >= mid_end:
        map_order = None

    layout = _ColumnLayout(
        src_library=src_library,
        src_dataset=src_dataset,
        src_variable=src_variable,
        src_label=src_label,
        map_definition=map_definition,
        map_rule=map_rule,
        map_mode=map_mode,
        map_order=map_order,
        sdtm_dataset=sdtm_dataset,
        sdtm_variable=sdtm_variable,
        sdtm_label=sdtm_label,
        sdtm_core=sdtm_core,
        total_cols=total_cols,
    )

    logger.debug("Column layout detected", layout=repr(layout))
    return layout


# ─────────────────────────────────────────────────────────────────────────────
# Safe Cell Access
# ─────────────────────────────────────────────────────────────────────────────

def _safe_cell(row: tuple, index: int | None) -> str:
    """
    Safely extract a cell value from a row tuple.

    Handles:
    - None index (column not present in this sheet)
    - Index out of bounds (row shorter than expected)
    - None cell value
    - Non-string cell values (numbers, dates)
    """
    if index is None or index < 0:
        return ""
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet Parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_single_sheet(
    ws,
    sheet_name: str,
    header_row_num: int = 2,
    data_start_row: int = 3,
) -> list[SpecMappingEntry]:
    """
    Parse one sheet from the spec workbook into SpecMappingEntry objects.

    Handles grouped-row format where Library and Dataset are only specified
    on the first row of a group and inherited by subsequent rows.
    """
    entries: list[SpecMappingEntry] = []

    # Read the header row to detect layout
    header_row = None
    for row in ws.iter_rows(
        min_row=header_row_num, max_row=header_row_num, values_only=True
    ):
        header_row = row
        break

    if header_row is None:
        logger.warning("Empty header row", sheet=sheet_name, row=header_row_num)
        return entries

    layout = _detect_layout(header_row)
    if layout is None:
        logger.warning("Could not detect column layout — skipping", sheet=sheet_name)
        return entries

    logger.info(
        "Layout detected",
        sheet=sheet_name,
        src_dataset_col=layout.src_dataset,
        sdtm_dataset_col=layout.sdtm_dataset,
        sdtm_core_col=layout.sdtm_core,
    )

    # State for inherited values (grouped row format)
    inherited_library = ""
    inherited_module = ""

    row_count = 0
    skip_count = 0

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        row_count += 1

        # Extract source fields
        library = _safe_cell(row, layout.src_library)
        src_dataset = _safe_cell(row, layout.src_dataset)
        src_variable = _safe_cell(row, layout.src_variable)
        src_label = _safe_cell(row, layout.src_label)

        # Extract mapping fields
        map_definition = _safe_cell(row, layout.map_definition)
        map_rule = _safe_cell(row, layout.map_rule)
        map_order = _safe_cell(row, layout.map_order)

        # Extract SDTM target fields
        sdtm_dataset = _safe_cell(row, layout.sdtm_dataset)
        sdtm_variable = _safe_cell(row, layout.sdtm_variable)
        sdtm_label = _safe_cell(row, layout.sdtm_label)
        sdtm_core = _safe_cell(row, layout.sdtm_core)

        # Handle inheritance (grouped row format)
        if library:
            inherited_library = library
        else:
            library = inherited_library

        if src_dataset:
            inherited_module = src_dataset
        else:
            src_dataset = inherited_module

        # Skip entirely empty rows
        if not sdtm_dataset and not sdtm_variable and not src_label and not map_definition:
            skip_count += 1
            continue

        # Skip rows with no SDTM target at all
        if not sdtm_dataset and not sdtm_variable:
            skip_count += 1
            continue

        # Normalize module name
        module_normalized = normalize_module_name(src_dataset)

        # Normalize label for exact lookup
        label_normalized = normalize_label_for_lookup(src_label)

        # Determine if supplemental
        sdtm_domain_upper = sdtm_dataset.upper() if sdtm_dataset else ""
        is_supp = sdtm_domain_upper.startswith("SUPP")

        # Normalize core value
        core_normalized = _normalize_core(sdtm_core)

        entry = SpecMappingEntry(
            library=library,
            source_module=module_normalized,
            source_variable=src_variable,
            source_label=src_label,
            source_label_normalized=label_normalized,
            map_definition=map_definition,
            map_rule=map_rule,
            map_order=map_order,
            sdtm_domain=sdtm_domain_upper,
            sdtm_variable=sdtm_variable,
            sdtm_label=sdtm_label,
            core=core_normalized,
            is_supplemental=is_supp,
        )
        entries.append(entry)

    logger.info(
        "Sheet parsed",
        sheet=sheet_name,
        data_rows=row_count,
        valid_entries=len(entries),
        skipped_rows=skip_count,
    )

    return entries


def _normalize_core(raw_core: str) -> str:
    """
    Normalize Core classification value.

    Handles variations: 'Req', 'Required', 'req', 'Perm', 'Permitted', etc.
    """
    if not raw_core:
        return ""

    core_lower = raw_core.lower().strip()

    if core_lower in ("req", "required"):
        return "Req"
    elif core_lower in ("perm", "permitted"):
        return "Perm"
    elif core_lower in ("cond", "conditional"):
        return "Cond"
    else:
        # Return as-is (might be empty or unknown)
        return raw_core.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Main Parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_az_spec(filepath: Path | None = None) -> tuple[list[SpecMappingEntry], SpecModuleIndex]:
    """
    Parse the complete AZ RAW-to-SDTM Specification workbook.

    Processes:
    1. Main 'RAW-SDTM Mappings' sheet (~24K rows)
    2. Additional dataset-specific sheets (BMRES2OM, PATHBIO, etc.)

    Skips non-mapping sheets (Changes, RELREC, etc.) as configured.

    Args:
        filepath: Path to the Excel file. Uses configured default if None.

    Returns:
        Tuple of (all_entries, module_index):
        - all_entries: Complete list of SpecMappingEntry objects
        - module_index: SpecModuleIndex for quick module lookups
    """
    filepath = filepath or AZ_SPEC_FILE

    if not filepath.exists():
        raise FileNotFoundError(f"AZ Specification file not found: {filepath}")

    logger.info("=" * 60)
    logger.info("Parsing AZ RAW-to-SDTM Specification", path=str(filepath))
    logger.info("=" * 60)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_entries: list[SpecMappingEntry] = []

    try:
        # ─── Main Sheet ───
        if EXCEL.spec_main_sheet not in wb.sheetnames:
            raise ValueError(
                f"Main sheet '{EXCEL.spec_main_sheet}' not found. "
                f"Available sheets: {wb.sheetnames}"
            )

        logger.info("Parsing main sheet", sheet=EXCEL.spec_main_sheet)
        main_entries = _parse_single_sheet(
            ws=wb[EXCEL.spec_main_sheet],
            sheet_name=EXCEL.spec_main_sheet,
            header_row_num=EXCEL.spec_header_row,
            data_start_row=EXCEL.spec_data_start_row,
        )
        all_entries.extend(main_entries)
        logger.info("Main sheet complete", entries=len(main_entries))

        # ─── Additional Dataset Sheets ───
        additional_sheets = [
            name for name in wb.sheetnames
            if name not in EXCEL.spec_skip_sheets
            and name != EXCEL.spec_main_sheet
        ]

        if additional_sheets:
            logger.info(
                "Processing additional sheets",
                count=len(additional_sheets),
                names=additional_sheets,
            )

        for sheet_name in additional_sheets:
            try:
                sheet_entries = _parse_single_sheet(
                    ws=wb[sheet_name],
                    sheet_name=sheet_name,
                    header_row_num=2,
                    data_start_row=3,
                )
                if sheet_entries:
                    all_entries.extend(sheet_entries)
                    logger.info(
                        "Additional sheet parsed",
                        sheet=sheet_name,
                        entries=len(sheet_entries),
                    )
            except Exception as e:
                logger.warning(
                    "Failed to parse additional sheet — skipping",
                    sheet=sheet_name,
                    error=str(e),
                )

    finally:
        wb.close()

    # ─── Apply Library Filter (if configured) ───
    if EXCEL.spec_library_filter:
        before_count = len(all_entries)
        all_entries = [
            e for e in all_entries
            if not e.library or any(
                f.lower() in e.library.lower()
                for f in EXCEL.spec_library_filter
            )
        ]
        logger.info(
            "Library filter applied",
            filter=EXCEL.spec_library_filter,
            before=before_count,
            after=len(all_entries),
        )

    # ─── Build Module Index ───
    module_index = _build_module_index(all_entries)

    # ─── Final Summary ───
    labeled_entries = sum(1 for e in all_entries if e.source_label.strip())
    supp_entries = sum(1 for e in all_entries if e.is_supplemental)

    logger.info("─" * 60)
    logger.info("AZ Spec parsing COMPLETE")
    logger.info(f"  Total mapping entries:    {len(all_entries):,}")
    logger.info(f"  Entries with labels:      {labeled_entries:,}")
    logger.info(f"  Supplemental entries:     {supp_entries:,}")
    logger.info(f"  Unique modules:           {len(module_index.module_to_domains)}")
    logger.info(f"  Unique target domains:    {len(set(e.sdtm_domain for e in all_entries if e.sdtm_domain))}")
    logger.info("─" * 60)

    return all_entries, module_index


def _build_module_index(entries: list[SpecMappingEntry]) -> SpecModuleIndex:
    """
    Build the module → domain index with primary domain detection.

    Primary domain = the most frequently occurring SDTM domain for that module.
    This is used for domain confidence in the resolution pipeline.
    """
    module_domains: dict[str, list[str]] = {}
    module_domain_counts: dict[str, Counter] = {}
    module_total_counts: dict[str, int] = {}
    module_label_counts: dict[str, int] = {}

    for entry in entries:
        module = entry.source_module
        if not module:
            continue

        if module not in module_domains:
            module_domains[module] = []
            module_domain_counts[module] = Counter()
            module_total_counts[module] = 0
            module_label_counts[module] = 0

        module_total_counts[module] += 1

        if entry.source_label.strip():
            module_label_counts[module] += 1

        if entry.sdtm_domain:
            module_domain_counts[module][entry.sdtm_domain] += 1

    # Resolve primary domain (most frequent) and sorted domain list
    module_to_primary: dict[str, str] = {}
    module_to_sorted_domains: dict[str, list[str]] = {}

    for module, counter in module_domain_counts.items():
        if counter:
            module_to_primary[module] = counter.most_common(1)[0][0]
            module_to_sorted_domains[module] = sorted(set(counter.keys()))
        else:
            module_to_primary[module] = ""
            module_to_sorted_domains[module] = []

    return SpecModuleIndex(
        module_to_domains=module_to_sorted_domains,
        module_to_primary_domain=module_to_primary,
        module_to_entry_count=module_total_counts,
        module_to_label_count=module_label_counts,
    )