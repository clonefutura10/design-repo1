"""
Parser for AZ Corporate SDTM Standards.

Extracts:
1. Datasets sheet (94 domains) — reference for domain validation
2. Variables sheet (4,535 entries) — source for CDISC FAISS Index 2
   AND validation reference for confirming variable existence

Column positions from inspection:
    Datasets: Dataset|Label|Layout|Level|Class|Description|Comment
    Variables: Dataset|Variable|Label|Type|Length|Core|Role|Format|Key|Seq|Origin|DatasetLevel|CDISCNotes|AZNotes
"""

from __future__ import annotations
from pathlib import Path

import openpyxl

from config.settings import AZ_CORPORATE_FILE, EXCEL
from src.cache_builder.models import CorporateDataset, CorporateVariable
from src.utils.logging_config import get_logger

logger = get_logger(__name__)


def _safe(row: tuple, idx: int) -> str:
    """Safely extract cell value from row tuple."""
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def parse_corporate_datasets(filepath: Path | None = None) -> list[CorporateDataset]:
    """
    Parse 'Datasets' sheet — domain definitions.

    Returns list of CorporateDataset (one per SDTM domain defined in AZ standards).
    """
    filepath = filepath or AZ_CORPORATE_FILE

    if not filepath.exists():
        raise FileNotFoundError(f"Corporate Standards file not found: {filepath}")

    logger.info("Parsing Corporate Standards — Datasets")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    try:
        ws = wb[EXCEL.corporate_datasets_sheet]
        datasets: list[CorporateDataset] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue

            dataset = _safe(row, 0)
            if not dataset:
                continue

            datasets.append(CorporateDataset(
                dataset=dataset,
                label=_safe(row, 1),
                layout=_safe(row, 2),
                level=_safe(row, 3),
                domain_class=_safe(row, 4),
                description=_safe(row, 5),
            ))

    finally:
        wb.close()

    logger.info("Corporate Datasets parsed", count=len(datasets))
    return datasets


def parse_corporate_variables(filepath: Path | None = None) -> list[CorporateVariable]:
    """
    Parse 'Variables' sheet — SDTM variable definitions.

    This is the source data for CDISC FAISS Index 2. Each variable's label
    and CDISC notes are embedded for semantic similarity search.

    Returns list of CorporateVariable (4,535 expected for current AZ standards).
    """
    filepath = filepath or AZ_CORPORATE_FILE

    if not filepath.exists():
        raise FileNotFoundError(f"Corporate Standards file not found: {filepath}")

    logger.info("Parsing Corporate Standards — Variables")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    try:
        ws = wb[EXCEL.corporate_variables_sheet]
        variables: list[CorporateVariable] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue

            dataset = _safe(row, 0)
            variable = _safe(row, 1)

            # Both dataset and variable are required
            if not dataset or not variable:
                continue

            variables.append(CorporateVariable(
                dataset=dataset,
                variable=variable,
                label=_safe(row, 2),
                variable_type=_safe(row, 3),
                length=_safe(row, 4),
                core=_safe(row, 5),
                role=_safe(row, 6),
                format_field=_safe(row, 7),
                key=_safe(row, 8),
                seq=_safe(row, 9),
                origin=_safe(row, 10),
                dataset_level=_safe(row, 11),
                cdisc_notes=_safe(row, 12) if len(row) > 12 else "",
                az_notes=_safe(row, 13) if len(row) > 13 else "",
            ))

    finally:
        wb.close()

    logger.info(
        "Corporate Variables parsed",
        count=len(variables),
        unique_datasets=len(set(v.dataset for v in variables)),
    )
    return variables