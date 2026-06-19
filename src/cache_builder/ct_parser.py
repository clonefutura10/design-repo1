"""
Parser for AZ SDTM Controlled Terminology (188K+ rows).

Deduplicates individual codelist values to unique variable-level associations.
Output: {"{domain}.{variable}": CTEntry} — used at annotation time to enrich
SDTM variable annotations with codelist references (e.g., "VSTEST (C66742)").
"""

from __future__ import annotations
from pathlib import Path
from collections import defaultdict

import openpyxl

from config.settings import AZ_CT_FILE, EXCEL
from src.cache_builder.models import CTEntry
from src.utils.logging_config import get_logger

logger = get_logger(__name__)


def parse_controlled_terminology(filepath: Path | None = None) -> dict[str, CTEntry]:
    """
    Parse AZ SDTM CT file to extract variable → codelist mappings.

    From 188K value-level rows, produces ~2-4K unique (domain, variable) → codelist
    associations. When a variable is associated with multiple codelists, the one
    with the most values is selected as primary.

    Column positions (from inspection):
        0: SDTM_DS (domain)
        1: SDTM_VAR (variable name)
        2: SDTM_LABEL (variable label)
        3: CODELIST (codelist identifier)
        4: CODELISTNAME (full codelist name)
        5: VALUE (individual codelist value — not used here)
        6: CODELISTCODE (NCI code for the codelist)
        7: VALUECODE (NCI code for the value — not used here)

    Returns:
        Dict keyed by "{domain}.{variable}" → CTEntry with codelist info.
    """
    filepath = filepath or AZ_CT_FILE

    if not filepath.exists():
        raise FileNotFoundError(f"CT file not found: {filepath}")

    logger.info("Parsing Controlled Terminology", path=str(filepath))

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    try:
        if EXCEL.ct_sheet not in wb.sheetnames:
            raise ValueError(
                f"CT sheet '{EXCEL.ct_sheet}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[EXCEL.ct_sheet]

        # Accumulate: key → {codelist → count of values}
        codelist_value_counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        # Store metadata for each codelist (last write wins for same codelist)
        codelist_metadata: dict[str, dict[str, dict]] = defaultdict(dict)

        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1

            # Bounds-safe access
            if not row or len(row) < 7:
                continue

            domain = str(row[0]).strip() if row[0] else ""
            variable = str(row[1]).strip() if row[1] else ""
            label = str(row[2]).strip() if row[2] else ""
            codelist = str(row[3]).strip() if row[3] else ""
            codelist_name = str(row[4]).strip() if row[4] else ""
            codelist_code = str(row[6]).strip() if row[6] else ""

            # Skip rows missing key fields
            if not domain or not variable or not codelist:
                continue

            key = f"{domain}.{variable}"
            codelist_value_counts[key][codelist] += 1
            codelist_metadata[key][codelist] = {
                "sdtm_domain": domain,
                "sdtm_variable": variable,
                "sdtm_label": label,
                "codelist": codelist,
                "codelist_name": codelist_name,
                "codelist_code": codelist_code,
            }

            # Progress logging every 50K rows
            if row_count % 50000 == 0:
                logger.info(f"  CT parsing progress: {row_count:,} rows processed")

    finally:
        wb.close()

    # For each (domain, variable), select the codelist with the most values
    ct_lookup: dict[str, CTEntry] = {}

    for key, counts in codelist_value_counts.items():
        primary_codelist = max(counts.keys(), key=lambda cl: counts[cl])
        info = codelist_metadata[key][primary_codelist]
        ct_lookup[key] = CTEntry(**info)

    logger.info(
        "CT parsing complete",
        total_rows_processed=row_count,
        unique_variable_mappings=len(ct_lookup),
        unique_domains=len(set(e.sdtm_domain for e in ct_lookup.values())),
    )

    return ct_lookup